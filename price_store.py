import os
from openpyxl import load_workbook

PRICES_XLSX = os.getenv("PRICES_XLSX", "/data/prices.xlsx")

def load_prices():
    """Return dict: { normalized_product_name: price_string } from Excel."""
    if not os.path.exists(PRICES_XLSX):
        return {}
    wb = load_workbook(PRICES_XLSX, data_only=True)
    if "Prices" not in wb.sheetnames:
        return {}
    ws = wb["Prices"]
    prices = {}
    # Expect header row: A1=product, B1=price
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # skip header
        if not row or (row[0] is None and row[1] is None):
            continue
        product = str(row[0]).strip() if row[0] is not None else ""
        price   = str(row[1]).strip() if row[1] is not None else ""
        if product:
            prices[product.lower()] = price
    return prices

def find_price_reply(message_text: str, prices_map: dict) -> str:
    """Very simple fuzzy contains: if any product name appears in the message, reply with its price."""
    if not message_text:
        return "Please tell me the product you want 😊"
    msg = message_text.lower()
    # Try longest product names first (more specific)
    for product in sorted(prices_map.keys(), key=len, reverse=True):
        if product in msg:
            price = prices_map[product]
            if price:
                return f"The price for {product.title()} is {price}."
            else:
                return f"I have {product.title()} but the price is not set yet."
    return "Please specify the exact product name (e.g., 'AirPods', 'iPhone 14', 'PS5')."
